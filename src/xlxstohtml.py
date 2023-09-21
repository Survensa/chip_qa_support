import openpyxl

def xlsx_to_html(input_file, output_file):
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(input_file)

    # Open the first sheet (you can modify this to select a specific sheet)
    sheet = workbook.active

    # Open the output file in write mode
    with open(output_file, 'w') as html_file:
        # Write the table opening tag
        html_file.write("<table>\n")

        # Write the table header
        headers = ["<th>" + (cell.value if cell.value is not None else "") + "</th>" for cell in next(sheet.iter_rows())]
        html_file.write("<tr>" + "".join(headers) + "</tr>\n")

        # Write the table rows
        for row in sheet.iter_rows(min_row=2):
            row_data = [("<td>" + str(cell.value) + "</td>") if cell.value is not None else "<td></td>" for cell in row]
            html_file.write("<tr>" + "".join(row_data) + "</tr>\n")

        # Write the table closing tag
        html_file.write("</table>\n")

if __name__ == "__main__":
    input_file = "Docs/TC_Summary.xlsx"  # Replace with your input Excel file
    output_file = "Docs/TC_Summary.html"  # Replace with your desired output HTML file name

    xlsx_to_html(input_file, output_file)
