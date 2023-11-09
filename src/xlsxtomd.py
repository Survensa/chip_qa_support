import openpyxl


def xlsx_to_md(input_file, output_file):
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(input_file)

    # Open the first sheet (you can modify this to select a specific sheet)
    sheet = workbook.active

    # Open the output file in write mode
    with open(output_file, "w") as md_file:
        # Write the table header
        headers = [
            "**" + cell.value + "**" if cell.value is not None else ""
            for cell in next(sheet.iter_rows())
        ]
        md_file.write("|".join(headers) + "|\n")
        md_file.write("|".join(["---"] * len(headers)) + "|\n")

        # Write the table rows
        for row in sheet.iter_rows(min_row=2):
            md_row = [str(cell.value) if cell.value is not None else "" for cell in row]
            md_file.write("|".join(md_row) + "|\n")


if __name__ == "__main__":
    input_file = "Docs/TC_Summary.xlsx"  # Replace with your input Excel file
    output_file = (
        "Docs/TC_Summary.md"  # Replace with your desired output Markdown file name
    )

    xlsx_to_md(input_file, output_file)
