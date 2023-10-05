from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import re
import json
import os
from datetime import datetime

print("Process Starts")

# Load existing JSON data or create an empty dictionary if it doesn't exist
json_filename = 'src/TC_Summary.json'
try:
    with open(json_filename, 'r') as json_file:
        existing_data = json.load(json_file)
except FileNotFoundError:
    existing_data = {}

# Define the filename
filename = 'Docs/TC_Summary.xlsx'

# Check if the file exists
if os.path.exists(filename):
    # Load the existing workbook
    workbook = openpyxl.load_workbook(filename)
    print(f"Existing workbook '{filename}' loaded.")
else:
    # Create a new workbook if the file doesn't exist
    workbook = openpyxl.Workbook()
    print(f"New workbook '{filename}' created.")

# Check if 'All_TC_Details' sheet exists, and if not, create it
sheet1_name = 'All_TC_Details'
if sheet1_name not in workbook.sheetnames:
    sheet1 = workbook.create_sheet(title=sheet1_name)

    # Define column headers
    headers = ['S.No', 'Cluster Name', 'Test Case Name', 'Test Case ID', 'Test Plan']

    # Add headers to the first row and set the font to bold for the headings
    header_font = Font(name='Times New Roman', bold=True)
    for col_num, header in enumerate(headers, 1):
        cell = sheet1.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = PatternFill(start_color='FFDDDDDD', end_color='FFDDDDDD', fill_type='solid')  # Light Grey Fill

    # Set header row alignment to center
    for cell in sheet1[1]:
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Set column width for S.No.
    sheet1.column_dimensions['A'].width = 5  # Set column width for S.No.

    print(f"Sheet '{sheet1_name}' created.")
else:
    # If the sheet exists, clear existing data
    sheet1 = workbook[sheet1_name]
    sheet1.delete_rows(2, sheet1.max_row)  # Clear existing data
    print(f"Sheet '{sheet1_name}' already exists. Existing data cleared.")

# Define a function to extract test case details
def extract_tc_details(h1_tags, a, row_number, sheet):
    for i, h1_tag in enumerate(h1_tags):
        h1 = h1_tag.text
        cluster_name = h1.replace(' Cluster Test Plan', '') \
            .replace(' Cluster TestPlan', '') \
            .replace(' Cluster', '') \
            .replace(' cluster', '') \
            .replace(' Test Plan', '')

        print("-" * 40)
        print(f"Fetching details for cluster: {cluster_name}")

        first_h1 = h1_tag
        if i == (len(h1_tags) - 1):
            second_h1 = False
        else:
            second_h1 = h1_tags[i + 1]

        h5_tags = first_h1.find_all_next('h5', {'id': lambda x: x and x.startswith('_test_procedure')})
        result = []

        if second_h1:
            for h5_tag in h5_tags:
                if h5_tag.find_previous('h1') == first_h1 and h5_tag.find_next('h1') == second_h1:
                    result.append(h5_tag)
        else:
            for h5_tag in h5_tags:
                if h5_tag.find_previous('h1') == first_h1:
                    result.append(h5_tag)

        if result:
            for j in range(len(result)):
                h4_tag = result[j].find_previous('h4')
                head_text_match = re.search(r'\[(.*?)\]\s*(.*)', h4_tag.text)
                if head_text_match:
                    head_text = '[' + head_text_match.group(1) + '] ' + head_text_match.group(2)
                else:
                    head_text = ''

                # Extract the "Test case name" using regular expressions
                testcase_match = re.search(r'\[(.*?)\]', head_text)
                if testcase_match:
                    testcase_name = testcase_match.group(1)  # Extract the first group (inside parentheses)
                else:
                    testcase_name = ''

                if a == 0:
                    test_plan = "Core Test Case"
                else:
                    test_plan = "App Test Case"

                # Modify row_values list to include "Test case name"
                row_values = [row_number, cluster_name, head_text, testcase_name, test_plan]
                sheet.append(row_values)

                print(f"Fetching details for Test Case: {testcase_name}")

                # Increment row_number for each new row
                row_number += 1

# Parse 'app' HTML
print("^" * 40)
print("Parsing 'app' HTML...")
app_html = 'Docs/Test_Plan_HTML/allclusters.html'
with open(app_html, encoding='utf-8') as f1:
    soup1 = BeautifulSoup(f1, 'html.parser')
    h1_tags1 = soup1.find_all('h1', {'id': True})
    extract_tc_details(h1_tags1, 1, 1, sheet1)  # Pass initial row_number and sheet1

# Calculate the next row_number after parsing the first HTML
row_number = sheet1.max_row + 0

# Check if 'TC_Changes' sheet exists, and if not, create it
changes_sheet_name = 'TC_Changes'
if changes_sheet_name not in workbook.sheetnames:
    changes_sheet = workbook.create_sheet(title=changes_sheet_name)

    changes_headers = ['Date of Run', 'Cluster Name', 'Test Case Name', 'Test Case ID', 'Test Plan', 'Change Type']
    # Add headers to the first row and set the font to bold for the headings
    for col_num, header in enumerate(changes_headers, 1):
        cell = changes_sheet.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = PatternFill(start_color='FFDDDDDD', end_color='FFDDDDDD', fill_type='solid')  # Light Grey Fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Set column widths for TC_Changes sheet
    column_widths = {'A': 30, 'B': 30, 'C': 100, 'D': 25, 'E': 15, 'F': 15}  # Adjusted widths for Date of Run and Change Type
    for column, width in column_widths.items():
        changes_sheet.column_dimensions[column].width = width

    # Set alignment to center for columns A, E, and F
    for column_letter in ['A', 'E', 'F']:
        for cell in changes_sheet[column_letter]:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    print(f"Sheet '{changes_sheet_name}' created.")
else:
    # If the sheet exists, do not clear existing data
    changes_sheet = workbook[changes_sheet_name]
    print(f"Sheet '{changes_sheet_name}' already exists.")

# Get the current date (with time)
current_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

# Add the added and removed test cases to the "TC_Changes" sheet
rows_to_insert = []

for cluster_name, cluster_data_list in added_test_cases.items():
    for cluster_data in cluster_data_list:
        rows_to_insert.append([current_date, cluster_name, cluster_data['Test Case Name'], cluster_data['Test Case ID'], cluster_data['Test Plan'], 'ADDED'])

for cluster_name, cluster_data_list in removed_test_cases.items():
    for cluster_data in cluster_data_list:
        rows_to_insert.append([current_date, cluster_name, cluster_data['Test Case Name'], cluster_data['Test Case ID'], cluster_data['Test Plan'], 'REMOVED'])

# Check if there are no changes and append a row indicating "No change"
if not rows_to_insert:
    changes_sheet.append([current_date, '-', '-', '-', '-', 'NO CHANGE'])

# Insert all rows at once
for row_values in rows_to_insert:
    changes_sheet.append(row_values)

# Set the font for the entire sheet to Times New Roman
for row in sheet1.iter_rows(min_row=2, max_row=sheet1.max_row, min_col=1, max_col=sheet1.max_column):
    for cell in row:
        cell.font = Font(name='Times New Roman')
        cell.alignment = Alignment(vertical='center')  # Center-align vertically

# Set alignment to center for columns A and E in sheet1
for column_letter in ['A', 'E']:
    for cell in sheet1[column_letter]:
        cell.alignment = Alignment(horizontal='center', vertical='center')

# Set column widths for sheet1
column_widths = {'A': 5, 'B': 30, 'C': 100, 'D': 25, 'E': 15}
for column, width in column_widths.items():
    sheet1.column_dimensions[column].width = width

# Save the workbook
print("Saving Excel workbook...")
workbook.save(filename)

# Update the JSON file with the latest data, excluding the history
with open(json_filename, 'w') as json_file:
    json.dump(current_data, json_file, indent=4)

print(f"Process completed. Excel file saved as '{filename}' and '{json_filename}' updated.")
