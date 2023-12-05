from bs4 import BeautifulSoup
import openpyxl
from openpyxl import load_workbook
from joblib import Parallel, delayed
import re
import datetime
import pandas as pd
import json

today_date = datetime.date.today().strftime('%Y-%m-%d')
excel_filename = "Docs/TC_Summary.xlsx"

try:
    workbook = load_workbook(excel_filename)
except FileNotFoundError:
    workbook = openpyxl.Workbook()

sheet_names = workbook.sheetnames

app_html_path = "Docs/Test_Plan_HTML/allclusters.html"
main_html_path = "Docs/Test_Plan_HTML/index.html"
json_filename = "src/TC_Summary.json"

def extract_testcase_id(testcase_header):
    match = re.search(r'\[(.*?)\]', testcase_header)
    if match:
        return match.group(1)
    return None

def extract_testcase_details(h4_tag, test_plan):
    details = {}
    header_text = h4_tag.text
    details["Test Case Name"] = header_text
    details["Test Case ID"] = extract_testcase_id(header_text)
    details["Test Plan"] = test_plan

    purpose_tag = h4_tag.find_next('h5', {'id': lambda x: x and x.startswith('_purpose')})
    purpose = purpose_tag.find_next('p').text
    details["Purpose"] = purpose

    details["PICS"] = []
    pics_tag = h4_tag.find_next('h5', {'id': lambda x: x and x.startswith('_pics')})
    ul_div = pics_tag.find_next('div', class_='ulist')
    p_tags = ul_div.find_all('p') if ul_div else []
    details["PICS"] = [p.text for p in p_tags]

    preconditions_tag = h4_tag.find_next('h5', {'id': lambda x: x and x.startswith('_preconditions')})
    if preconditions_tag:
        details["Pre-condition"] = {}
        table = preconditions_tag.find_next('table')
        if table:
            data_dict = create_dataframe_from_table(table)
            details["Pre-condition"] = data_dict
    else:
        details["Pre-condition"] = "Nil"

    procedure_tags = h4_tag.find_next('h5', {'id': lambda x: x and x.startswith('_test_procedure')}) 
    if not procedure_tags:
        procedure_tags = h4_tag.find_next('h6', {'id': lambda x: x and x.startswith('_test_procedure')}) 
    table = procedure_tags.find_next('table')
    data_dict = create_dataframe_from_table(table)
    details["Test Procedure"] = data_dict

    workbook.save(excel_filename)

    return header_text, details

def extract_testcase_details_from_html(header_tag, is_app, second_h1_tag=None):
    if is_app:
        html_path = app_html_path
    else:
        html_path = main_html_path

    with open(html_path) as html_file:
        soup = BeautifulSoup(html_file, 'lxml')

    h1_tags = soup.find_all('h1', {'id': True})
    first_h1_tag = None

    for tag in h1_tags:
        tag_text = tag.text
        if tag_text == header_tag:
            first_h1_tag = tag
            break

    if not first_h1_tag:
        return None

    if second_h1_tag:
        for tag in h1_tags:
            tag_text = tag.text
            if tag_text == second_h1_tag:
                second_h1 = tag
                break
    else:
        second_h1 = None

    h1 = first_h1_tag.text
    cluster_name = h1.replace(' Cluster Test Plan', '') \
        .replace(' Cluster TestPlan', '') \
        .replace(' Cluster', '') \
        .replace(' cluster', '') \
        .replace(' Test Plan', '')

    if h1 == "MCORE PICS Definition":
        return None

    clus = []
    first_h1 = first_h1_tag

    h5_tags = first_h1.find_all_next(['h5', 'h6'], {'id': lambda x: x and x.startswith('_test_procedure')})

    result = []
    if second_h1:
        for h5_tag in h5_tags:
            if h5_tag.find_previous('h1') == first_h1:
                if h5_tag.find_next('h1') == second_h1:
                    result.append(h5_tag)
    else:
        for h5_tag in h5_tags:
            result.append(h5_tag)

    if result:
        heads = []
        for h5_tag in result:
            h4_tag = h5_tag.find_previous('h4')
            headt, d = tc(h4_tag, "App Test Case" if is_app else "Core Test Case")
            clus.append(d)
            heads.append(headt)
    else:
        h5_tags = first_h1.find_all_next(['h5', 'h6'], {'id': lambda x: x and x.startswith('_test_procedure')})

        for h5_tag in h5_tags:
            if h5_tag.find_previous('h1') == first_h1:
                if h5_tag.find_next('h1') == second_h1:
                    result.append(h5_tag)

        heads = []
        for h5_tag in result:
            h4_tag = h5_tag.find_previous('h5')
            headt, d = tc(h4_tag, "App Test Case" if is_app else "Core Test Case")
            clus.append(d)
            heads.append(headt)

    return {cluster_name: clus}

def create_dataframe_from_table(table_tag):
    rows = table_tag.find_all('tr')
    data = []
    col_spans = []
    row_spans = []

    for i, row in enumerate(rows):
        cells = row.find_all(['th', 'td'])
        row_data = []
        for j, cell in enumerate(cells):
            if cell.has_attr('colspan'):
                col_spans.append((i, j, int(cell['colspan'])))
            if cell.has_attr('rowspan'):
                row_spans.append((i, j, int(cell['rowspan'])))
            row_data.append(cell.get_text(strip=True))
        data.append(row_data)

    for span in col_spans:
        i, j, span_size = span
        for r in range(1, span_size):
            data[i].insert(j + 1, '')

    for span in row_spans:
        i, j, span_size = span
        for r in range(1, span_size):
            data[i + r].insert(j, '')

    df = pd.DataFrame(data[1:], columns=data[0])
    df = df.fillna('')
    data_dict = df.to_dict('list')
    return data_dict

def compare_existing_updated_data(existing_data, updated_data):
    new_clusters = list(updated_data.keys())
    old_clusters = list(existing_data.keys())

    added_clusters = list(set(new_clusters).difference(set(old_clusters)))
    removed_clusters = list(set(old_clusters).difference(set(new_clusters)))
    changed_testcases = {}
    new_testcases = []
    removed_testcases = []

    for cluster in new_clusters:
        if cluster in added_clusters + removed_clusters:
            continue
        new_testcases_data = updated_data[cluster]
        old_testcases_data = existing_data.get(cluster, [])

        if new_testcases_data == old_testcases_data:
            print(f"No changes in {cluster} cluster")
            continue

        if len(new_testcases_data) == len(old_testcases_data):
            for i in range(len(new_testcases_data)):
                if new_testcases_data[i] == old_testcases_data[i]:
                    print(f"{new_testcases_data[i]['Test Case ID']} has no change ")
                else:
                    changes = []
                    keys = list(new_testcases_data[i].keys())
                    for key in keys:
                        if new_testcases_data[i][key] == old_testcases_data[i][key]:
                            print(f"{new_testcases_data[i]['Test Case ID']} {key} has no change ")
                        elif key == "Test Procedure":
                            procedure_keys = list(new_testcases_data[i][key].keys())
                            print(procedure_keys)
                            for proc_key in procedure_keys:
                                if new_testcases_data[i][key][proc_key] == old_testcases_data[i][key][proc_key]:
                                    print(f"{new_testcases_data[i]['Test Case ID']} {proc_key} has no change ")
                                else:
                                    changes.append(f"Testprocedure({proc_key})")
                        else:
                            changes.append(key)
                    changed_testcases[new_testcases_data[i]['Test Case ID']] = changes

        elif len(new_testcases_data) > len(old_testcases_data):
            new_clusters.append(cluster)

        else:
            removed_testcases.append(cluster)

    differences = {
        "added_clusters": added_clusters,
        "removed_clusters": removed_clusters,
        "changed_testcases": changed_testcases,
        "added_testcases": new_testcases,
        "removed_testcases": removed_testcases
    }

    return differences

def update_change_log(differences, sheet, version):
    changes = []
  
    if differences["changed_testcases"]:
        keys = list(differences["changed_testcases"].keys())
        for key in keys:
            for change in differences["changed_testcases"][key]:
                changes.append([today_date, version, key, "Testcase Modified", change])

    if differences["added_clusters"]:
        for cluster in differences["added_clusters"]:
            changes.append([today_date, version, cluster, "Cluster Added"])

    if differences["removed_clusters"]:
        for cluster in differences["removed_clusters"]:
            changes.append([today_date, version, cluster, "Cluster Removed"])

    if differences["added_testcases"]:
        for testcase in differences["added_testcases"]:
            changes.append([today_date, version, testcase, "Testcase Added"])

    if differences["removed_testcases"]:
        for testcase in differences["removed_testcases"]:
            changes.append([today_date, version, testcase, "Testcase Removed"])

    if changes:
        print(f"Number of Testcase Change : {len(changes)}\n")
        print(changes)
        for _ in range(len(changes)):
            sheet.insert_rows(2)
        for i, change_data in enumerate(changes):
            for j, value in enumerate(change_data):                                                                                 
                sheet.cell(row=i + 2, column=j + 1, value=value)
    else:
        sheet.insert_rows(2)
        default_value = [today_date, version, "Nil", f"NO CHANGE : {today_date}", "Nil"]
        for _ in range(1):
            for j, value in enumerate(default_value):                                                                                 
                sheet.cell(row=_ + 2, column=j + 1, value=value)
    
    return None

def update_testcase_change_log(differences, sheet, version):
    changes = []
  
    if differences["changed_testcases"]:
        keys = list(differences["changed_testcases"].keys())
        for key in keys:
            for change in differences["changed_testcases"][key]:
                if change in ["Test Case Name", "Test Case ID"]:
                    changes.append([today_date, version, key, "Testcase Modified", change])

    if differences["added_clusters"]:
        for cluster in differences["added_clusters"]:
            changes.append([today_date, version, cluster, "Cluster Added"])

    if differences["removed_clusters"]:
        for cluster in differences["removed_clusters"]:
            changes.append([today_date, version, cluster, "Cluster Removed"])

    if differences["added_testcases"]:
        for testcase in differences["added_testcases"]:
            changes.append([today_date, version, testcase, "Testcase Added"])

    if differences["removed_testcases"]:
        for testcase in differences["removed_testcases"]:
            changes.append([today_date, version, testcase, "Testcase Removed"])
		
    if changes:
        print(f"Number of Testcase Change : {len(changes)}\n")
        print(changes)
        for _ in range(len(changes)):
            sheet.insert_rows(2)
        for i, change_data in enumerate(changes):
            for j, value in enumerate(change_data):                                                                                 
                sheet.cell(row=i + 2, column=j + 1, value=value)
    else:
        sheet.insert_rows(2)
        default_value = [today_date, version, "Nil", f"NO CHANGE : {today_date}", "Nil"]
        for _ in range(1):
            for j, value in enumerate(default_value):                                                                                 
                sheet.cell(row=_ + 2, column=j + 1, value=value)

    return None

def create_cluster_enclosure_tags(h1_tags):
    enclosure_tags = []
    for i in range(len(h1_tags)):
        if i == (len(h1_tags) - 1):
            second_h1_tag = False
        else:
            second_h1_tag = h1_tags[i + 1]
        enclosure_tags.append(second_h1_tag)
    return enclosure_tags

def update_testcase_summary(sheet, updated_data):
    print("Updating test case summary in Excel sheet 'All_TC_Details'...")
    clusters = list(updated_data.keys())
    for cluster in clusters:
        tc_ids = updated_data[cluster][0]["Test Case ID"]
        sh = re.search(r'-(.*?)-', tc_ids)
        code = sh.group(1)
        if code == 'LOWPOWER':
            code = 'MC'
        sheet_name = workbook[code]
        sheet_name.append([cluster])
        sheet_name.append([""])

        testcases_data = updated_data[cluster]
        for testcase_data in testcases_data:
            testcase_name = testcase_data["Test Case Name"]
            sheet_name.append([testcase_name])
            sheet_name.append([""])
            sheet_name.append(["Purpose"])
            sheet_name.append([testcase_data["Purpose"]])
            sheet_name.append([""])
            sheet_name.append(["PICS"])
            for pics in testcase_data["PICS"]:
                sheet_name.append([pics])
            sheet_name.append([""])
            sheet_name.append(["Pre-condition"])
            if testcase_data["Pre-condition"] == "Nil":
                sheet_name.append(["Nil"])
            else:
                head = list(testcase_data["Pre-condition"].keys())
                sheet_name.append(head)
                for i in range(len(list(testcase_data["Pre-condition"].values())[0])):
                    values = []
                    for key, value in testcase_data["Pre-condition"].items():
                        if i < len(value):
                            values.append(value[i]) 
                    sheet_name.append(values)
            sheet_name.append([""])
            sheet_name.append(["Test Procedure"])
            keys = list(testcase_data["Test Procedure"].keys())
            sheet_name.append(keys)
            for i in range(len(list(testcase_data["Test Procedure"].values())[0])):
                values = []
                for key, value in testcase_data["Test Procedure"].items():
                    if i < len(value):
                        values.append(value[i])
                sheet_name.append(values)
            sheet_name.append([""])
            sheet_name.append([""])
            sheet_name.append([""])
            head_text_match = re.search(r'\[(.*?)\]\s*(.*)', testcase_name)
            if head_text_match:
                tc_name = '[' + head_text_match.group(1) + '] ' + head_text_match.group(2)
            else:
                tc_name = ''
            tc_id = testcase_data["Test Case ID"]
            test_plan = testcase_data["Test Plan"]
            row_number = sheet.max_row
            row_data = [row_number, cluster, tc_name, tc_id, test_plan]
            sheet.append(row_data)
            print(f"Updated details for test case: {test_case['Test Case ID']} in cluster: {cluster}")
    workbook.save(excel_filename)
    print("Test case summary updated successfully!")

if __name__ == '__main__':
    print("Starting the script...")
    updated_test_cases = {}

    try:
        with open(json_filename, 'r') as json_file:
            existing_test_cases = json.load(json_file)
            is_existing_data_present = True
    except FileNotFoundError:
        existing_test_cases = {}
        is_existing_data_present = False

    with open(app, 'r') as app_file:
        app_soup = BeautifulSoup(app_file, 'lxml')

    with open(main, 'r') as main_file:
        main_soup = BeautifulSoup(main_file, 'lxml')

    version_tag = app_soup.find('div', class_='details')
    version_text = version_tag.find('span', id="revnumber")
    version_number = version_text.text

    app_h1_tags = app_soup.find_all('h1', {'id': True})
    main_h1_tags = main_soup.find_all('h1', {'id': True})

    app_h1_tags_text = []
    main_h1_tags_text = []

    for tag in app_h1_tags:
        tag_text = tag.text
        app_h1_tags_text.append(tag_text)

    for tag in main_h1_tags:
        tag_text = tag.text
        main_h1_tags_text.append(tag_text)

    print(f"Number of clusters in app HTML: {len(app_h1_tags)}")
    print(f"Number of clusters in app HTML: {len(main_h1_tags)}")
	
    if "All_TC_Details" in sheet_names:
        workbook.remove(workbook["All_TC_Details"])
        workbook.create_sheet("All_TC_Details", 0)
        all_tc_details_sheet = workbook["All_TC_Details"]
    else:
        all_tc_details_sheet = workbook.active
        all_tc_details_sheet.title = "All_TC_Details"

    header_row = ["S.no", "Cluster Name", "Test Case Name", "Test Case ID", " Test Plan "]
    all_tc_details_sheet.append(header_row)

    app_enclosure_tags = cluster_enclose(app_h1_tags)
    print(f"Cluster enclosures in app HTML: {app_enclosure_tags}")

    if len(app_enclosure_tags) == len(app_h1_tags):
        input_data = []
        for i in range(len(app_enclosure_tags)):
            current_h1_tag = app_h1_tags_text[i]
            is_app_test_case = 0
            current_enclosure_tag = app_enclosure_tags[i]
            input_data.append((current_h1_tag, is_app_test_case, current_enclosure_tag))
	
	print("Extracting test case details from app HTML...")
        results = Parallel(n_jobs=-1)(delayed(tc_details)(a, b, c) for a, b, c in input_data)
	    
	for result in results:
            if result is not None:
                updated_test_cases.update(result)
    else:
        print("Failed to extract test case details from app HTML")

    main_enclosure_tags = cluster_enclose(main_h1_tags)
    print(f"Cluster enclosures in app HTML: {main_enclosure_tags}")

    if len(main_enclosure_tags) == len(main_h1_tags):
        input_data = []
        for i in range(len(main_enclosure_tags)):
            current_h1_tag = main_h1_tags_text[i]
            is_app_test_case = 1
            current_enclosure_tag = main_enclosure_tags[i]
            input_data.append((current_h1_tag, is_app_test_case, current_enclosure_tag))
	
	print("Extracting test case details from main HTML...")
        results = Parallel(n_jobs=-1)(delayed(tc_details)(a, b, c) for a, b, c in input_data)

        for result in results:
            if result is not None:
                updated_test_cases.update(result)
    else:
        print("Failed to extract test case details from main HTML")

    test_case_codes = []
    existing_clusters = list(existing_test_cases.keys())

    for test_case_name in updated_test_cases:
        test_case_id = updated_test_cases[test_case_name][0]["Test Case ID"]
        search_result = re.search(r'-(.*?)-', test_case_id)
        test_case_code = search_result.group(1)
        if test_case_code == 'LOWPOWER':
            test_case_code = 'MC'

        test_case_codes.append(test_case_code)

    for code in test_case_codes:
        if code in sheet_names:
            workbook.remove(workbook[code])
            workbook.create_sheet(code, sheet_names.index(code))
        else:
            workbook.create_sheet(code)
			
	print("Updating test case summary in Excel...")
    update_all_test_case_details(all_tc_details_sheet, updated_test_cases)

    if is_existing_data_present:
	print("Detecting changes in test cases...")
        differences = identify_changes(existing_test_cases, updated_test_cases)
	print("\nChanges detected:")
        print(differences)

        if "Test_plan_Changes" not in sheet_names:
            change_log_sheet = workbook.create_sheet("Test_plan_Changes", 2)
            change_log_sheet.append(["Date of Run", " Commit", "Cluster/Testcase", "Changes", "Column"])
        else:
            change_log_sheet = workbook["Test_plan_Changes"]
		
	print("\nUpdating Test Plan Changes sheet...")
        update_change_log_sheet(differences, change_log_sheet, version_number)

        print("\nChanges detected:")
        print(differences)

        if "Test_Summary_Changes" not in sheet_names:
            test_summary_changes_sheet = workbook.create_sheet("Test_Summary_Changes", 1)
            test_summary_changes_sheet.append(
                ["Date of Run", "Cluster Name", "Test Case Name", "Test Case ID", "Test Plan", "Change Type"]
            )
        else:
            test_summary_changes_sheet = workbook["Test_Summary_Changes"]
		
	print("\nUpdating Test Summary Changes sheet...")
        update_test_summary_changes_sheet(differences, test_summary_changes_sheet, version_number)

    column_widths = {'A': 10, 'B': 20, 'C': 50, 'D': 30, 'E': 30}  # Specify the column widths as desired

    for column, width in column_widths.items():
        all_tc_details_sheet.column_dimensions[column].width = width

    workbook.save(filename)
    print("Saving updated data to JSON file...")

    with open(json_filename, 'w') as json_file:
        json.dump(updated_test_cases, json_file, indent=4)

    workbook.save(filename)
    print("\nScript execution completed.")
