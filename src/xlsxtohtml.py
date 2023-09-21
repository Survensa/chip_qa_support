import openpyxl

def xlsx_to_html(input_file, output_file):
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(input_file)

    # Open the first sheet (you can modify this to select a specific sheet)
    sheet = workbook.active

    # Open the output file in write mode
    with open(output_file, 'w') as html_file:
        # Write the table opening tag with CSS styling for borders and header background color
        html_file.write('<table style="border-collapse: collapse; width: 100%;">\n')

        # Write the table header with a light gray background color
        headers = ["<th style='border: 1px solid black; padding: 8px; background-color: #f2f2f2;'>" + (cell.value if cell.value is not None else "") + "</th>" for cell in next(sheet.iter_rows())]
        html_file.write("<tr>" + "".join(headers) + "</tr>\n")

        # Write the table rows
        for row in sheet.iter_rows(min_row=2):
            row_data = [("<td style='border: 1px solid black; padding: 8px;'>" + str(cell.value) + "</td>") if cell.value is not None else "<td></td>" for cell in row]
            html_file.write("<tr>" + "".join(row_data) + "</tr>\n")

        # Write the table closing tag
        html_file.write("</table>\n")

if __name__ == "__main__":
    input_file = "Docs/TC_Summary.xlsx"  # Replace with your input Excel file
    output_file = "Docs/TC_Summary.html"  # Replace with your desired output HTML file name

    xlsx_to_html(input_file, output_file)
